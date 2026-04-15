from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import sqlite3
import uuid


@dataclass
class ImportDiagnostic:
    correlation_id: str
    parser: str
    ocr_engine: str | None
    extracted_chars: int
    status: str
    diagnostic: str


class ImportService:
    def __init__(self, conn: sqlite3.Connection):
        self.conn = conn

    @staticmethod
    def _ocr_pdf(path: Path) -> tuple[str, str]:
        try:
            from pdf2image import convert_from_path  # type: ignore
            import pytesseract  # type: ignore
        except Exception as exc:
            return "", f"OCR unavailable: {exc}"

        try:
            pages = convert_from_path(str(path), dpi=300)
            extracted: list[str] = []
            for page in pages:
                extracted.append(pytesseract.image_to_string(page, lang="eng+rus"))
            text = "\n".join(extracted)
            return text, f"OCR completed on {len(pages)} pages"
        except Exception as exc:
            return "", f"OCR failed: {exc}"

    @staticmethod
    def _extract_text(path: Path) -> tuple[str, str, str | None, str]:
        ext = path.suffix.lower()

        if ext in {".txt", ".csv"}:
            return path.read_text(encoding="utf-8", errors="ignore"), "plain", None, "Текст извлечен"

        if ext == ".pdf":
            try:
                from pypdf import PdfReader  # type: ignore

                reader = PdfReader(str(path))
                text = "\n".join(page.extract_text() or "" for page in reader.pages)
                if text.strip():
                    return text, "pypdf", None, "Текст извлечен из PDF"

                ocr_text, ocr_diag = ImportService._ocr_pdf(path)
                if ocr_text.strip():
                    return ocr_text, "pypdf+ocr", "tesseract", ocr_diag
                return "", "pypdf+ocr", "tesseract", ocr_diag
            except Exception as exc:
                return "", "pdf_failed", None, f"Ошибка PDF парсинга: {exc}"

        if ext in {".xlsx", ".xls"}:
            try:
                from openpyxl import load_workbook  # type: ignore

                workbook = load_workbook(path, read_only=True, data_only=True)
                lines: list[str] = []
                for ws in workbook.worksheets:
                    for row in ws.iter_rows(values_only=True):
                        vals = [str(v).strip() for v in row if v is not None and str(v).strip()]
                        if vals:
                            lines.append(" | ".join(vals))
                text = "\n".join(lines)
                return text, "openpyxl", None, "Текст извлечен из Excel"
            except Exception as exc:
                return "", "excel_failed", None, f"Ошибка Excel парсинга: {exc}"

        if ext in {".doc", ".docx"}:
            try:
                from docx import Document  # type: ignore

                doc = Document(str(path))
                text = "\n".join(p.text for p in doc.paragraphs)
                return text, "python-docx", None, "Текст извлечен из Word"
            except Exception as exc:
                return "", "word_failed", None, f"Ошибка Word парсинга: {exc}"

        return "", "unsupported", None, f"Неподдерживаемый формат: {ext}"

    def import_file(self, path: Path) -> ImportDiagnostic:
        correlation_id = uuid.uuid4().hex
        ext = path.suffix.lower()

        try:
            text, parser, ocr_engine, diagnostic = self._extract_text(path)

            if parser == "unsupported":
                status = "failed"
            elif text.strip():
                status = "success"
            elif "OCR" in diagnostic:
                status = "needs_review"
            elif parser.endswith("_failed"):
                status = "failed"
            else:
                status = "empty"

            run = ImportDiagnostic(
                correlation_id=correlation_id,
                parser=parser,
                ocr_engine=ocr_engine,
                extracted_chars=len(text),
                status=status,
                diagnostic=diagnostic,
            )
            self.conn.execute(
                """
                INSERT INTO import_runs(correlation_id, source_filename, source_type, parser, ocr_engine, ocr_pages, extracted_chars, status, diagnostic)
                VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    run.correlation_id,
                    path.name,
                    ext or "unknown",
                    run.parser,
                    run.ocr_engine,
                    0,
                    run.extracted_chars,
                    run.status,
                    run.diagnostic,
                ),
            )
            self.conn.commit()
            return run
        except Exception as exc:
            fail = ImportDiagnostic(
                correlation_id=correlation_id,
                parser="import_exception",
                ocr_engine=None,
                extracted_chars=0,
                status="failed",
                diagnostic=f"Ошибка импорта: {exc}",
            )
            self.conn.execute(
                """
                INSERT INTO import_runs(correlation_id, source_filename, source_type, parser, ocr_engine, ocr_pages, extracted_chars, status, diagnostic)
                VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    fail.correlation_id,
                    path.name,
                    ext or "unknown",
                    fail.parser,
                    fail.ocr_engine,
                    0,
                    fail.extracted_chars,
                    fail.status,
                    fail.diagnostic,
                ),
            )
            self.conn.commit()
            return fail
